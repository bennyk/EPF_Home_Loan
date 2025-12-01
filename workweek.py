import pandas as pd
import holidays
from datetime import timedelta, date

from holidays import country_holidays
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# Color palette, global vars
hex_color = "DAEEF3"
pattern_fill: PatternFill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

hex_color2 = "B7DEE8"
pattern_fill2: PatternFill = PatternFill(start_color=hex_color2, end_color=hex_color2, fill_type="solid")

hex_color3 = "D8E4BC"
pattern_fill3 = PatternFill(start_color=hex_color3, end_color=hex_color3, fill_type="solid")

def colnum_alphabet(n):
    alpha = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        alpha = chr(65 + remainder) + alpha
    return alpha

def iso_year_start(year: int) -> date:
    jan4   = date(year,  1, 4)
    return jan4 - timedelta(days=jan4.isoweekday() - 1)

def iso_year_end(year: int) -> date:
    return iso_year_start(year + 1) - timedelta(days=1)

def consume_shift(sheet: Worksheet, i: int, j: int):
    """
    Applies formatting and header values to a 3-column block (Shift, OT, Total).

    The block starts at column index j + i + 3 and spans rows 1, 2, and 3.
    """

    # Define the starting column index for the "Shift" block
    start_col = j + i + 3

    # Define the header labels for the three columns
    headers = {0: "Shift", 1: "OT", 2: "Total"}

    # 1. Apply fill formatting to the 3-column block (Rows 1 and 2)
    # The block is defined by: Rows (1 to 2) and Columns (start_col to start_col + 2)
    for r in range(1, 3):
        for c_offset in range(3):
            col = start_col + c_offset
            sheet.cell(row=r, column=col).fill = pattern_fill

    # 2. Apply headers and fill formatting to Row 3
    for c_offset in range(3):
        col = start_col + c_offset
        cell = sheet.cell(row=3, column=col)

        # Set the header value
        cell.value = headers.get(c_offset, "")

        # Apply the fill
        cell.fill = pattern_fill

    # 3. Set the column widths
    for c_offset in range(3):
        col = start_col + c_offset
        col_letter = colnum_alphabet(col)
        sheet.column_dimensions[col_letter].width = 5


class CalendarBuilder:
    def __init__(self, shift_config=None, employees: list=None):
        self.wb = None
        self.start_row = 5
        self.max_shifts = 2
        if shift_config is not None:
            self.max_shifts = len(shift_config)

        # Generate employee names (A to H for 2 shifts * 4 employees)
        self.employees_per_shift = 8
        self.total_employees = self.max_shifts * self.employees_per_shift
        if employees is not None:
            self.total_employees = len(employees)
        self.employees_per_shift = self.total_employees // self.max_shifts

        # TODO self.employee_names = [colnum_alphabet(i) for i in range(1, self.total_employees + 1)]
        self.employee_names = employees.copy()

        # Specify the country for which to check holidays (e.g., United States)
        # You can replace 'US' with the code for your desired country (e.g., 'GB' for United Kingdom, 'DE' for Germany)
        self.my_holidays = holidays.country_holidays('MY', subdiv='PNG', years=2025)

    def add_public_holidays(self, sheet, full_weeks):
        for shift in range(self.max_shifts):
            for emp_idx in range(self.employees_per_shift):
                for i in range(len(full_weeks)):
                    row = self.start_row + shift * (self.employees_per_shift + 1) + emp_idx
                    current_date = full_weeks['Date'].iloc[i].date()
                    if current_date in self.my_holidays:
                        # If Public Holiday falls on Sat/Sunday in Malaysia, example PH on Merdeka falls on Sunday, WW35 2025
                        #  "Under the Employment Act, if a public holiday falls on a rest day (such as Saturday and Sunday),
                        #   the next working day (Monday) will be allocated as paid compensatory leave."
                        sheet.cell(row=row, column=i+2).value = "PH"

                        # Loop through all holidays in the year and check if they are in the target month
                        holiday_name = self.my_holidays.get(current_date)
                        note_text = holiday_name
                        # TODO hmm the note_author is omitted
                        note_author = "Python Script"
                        comment = Comment(note_text, note_author)
                        sheet.cell(row=row, column=i+2).comment = comment

                        if full_weeks['Weekday'].iloc[i] in ("Saturday", "Sunday", "Sat", "Sun"):
                            note_text = "Replacement holiday"
                            note_author = "Python Script"
                            comment = Comment(note_text, note_author)
                            days_limit = 28
                            if current_date.day+2 <= days_limit:
                                sheet.cell(row=row, column=i+4).value = "PH*"
                                sheet.cell(row=row, column=i+4).comment = comment
                            else:
                                if current_date.day+1 <= days_limit:
                                    sheet.cell(row=row, column=i+3).value = "PH*"
                                    sheet.cell(row=row, column=i+3).comment = comment
                                else:
                                    # TODO Employee "A" miscalculated to Night Shift on August 31. He should be employee "E"
                                    # A 1 2025-08-31 34: Index out of bound
                                    print(f"{self.employee_names[emp_idx]} {shift} {current_date} {i}: Index out of bound")

        # for date, name in sorted(my_holidays.items()):
        #     print(f"{date}: {name}")
        #     sheet.cell(row=1, column=i+2).value = row['ISO_Week']

    def add_public_holidays2(self, sheets: list[Worksheet], df: pd.DataFrame):
        for shift in range(self.max_shifts):
            for emp_idx in range(self.employees_per_shift):
                row = self.start_row + shift * (self.employees_per_shift + 1) + emp_idx
                for i in range(len(df)):
                    current_date = df['Date'].iloc[i].date()
                    if current_date in self.my_holidays:
                        holiday_name = self.my_holidays.get(current_date)
                        # TODO Looks like the replacement holiday has been rescheduled automatically by Holidays Python module
                        if df['Weekday'].iloc[i] in ("Saturday", "Sunday", "Sat", "Sun"):
                            if current_date.day >= 30:
                                print(emp_idx, current_date, holiday_name)
                                month = df['Date'].iloc[i+1].date().month
                                sheets[month-1].cell(row=row, column=2).value = "PH*"
                                comment = Comment("Replacement holiday", "Python script")
                                sheets[month-1].cell(row=row, column=2).comment = comment

    def add_sample_employees(self, sheet: Worksheet, full_weeks: pd.DataFrame) -> None:
        """
        Populate an Excel sheet with sample employee names (A to H) in the first column.

        Args:
            sheet: Worksheet object (e.g., openpyxl Worksheet)
            full_weeks: Dictionary of full weeks indexed by date.
        """
        # Write employee names and schedule
        employee_index = 0

        # Generate legends
        last_row = self.start_row + (self.max_shifts - 1) * (self.employees_per_shift + 1) * 2
        sheet.cell(row=last_row, column=1).value = "Legends:"
        # TODO Assuming Shift is mandatory
        sheet.cell(row=last_row, column=1).font = Font(bold=True)

        sheet.cell(row=last_row+1, column=1).value = "  Present"
        sheet.cell(row=last_row+1, column=2).value = "P"
        sheet.cell(row=last_row+2, column=1).value = "  Public Holiday"
        sheet.cell(row=last_row+2, column=2).value = "PH"
        sheet.cell(row=last_row+3, column=1).value = "  Overtime"
        sheet.cell(row=last_row+3, column=2).value = "OT"
        sheet.cell(row=last_row+4, column=1).value = "  Half day"
        sheet.cell(row=last_row+4, column=2).value = "HD"
        sheet.cell(row=last_row+5, column=1).value = "  Medical cert"
        sheet.cell(row=last_row+5, column=2).value = "MC"
        sheet.cell(row=last_row+6, column=1).value = "  Training"
        sheet.cell(row=last_row+6, column=2).value = "T"
        sheet.cell(row=last_row+7, column=1).value = "  Annual leave"
        sheet.cell(row=last_row+7, column=2).value = "AL"
        sheet.cell(row=last_row+8, column=1).value = "  Holiday"
        sheet.cell(row=last_row+8, column=2).value = "H"

        # Add a list of options for worksheet
        data_val = DataValidation(type="list", formula1=f'=$B${last_row+1}:$B${last_row+8}', allowBlank=True)
        sheet.add_data_validation(data_val)

        # Shift change, Olive Green

        # TODO Consecutive two month working days, end of March to beginning of April
        shift_pattern = 0

        # Write employee names to the first column
        for shift in range(self.max_shifts):
            # print(employee_index)
            start = self.start_row + shift * (self.employees_per_shift + 1)
            sheet.row_dimensions[start-1].height = 10
            for emp_idx in range(self.employees_per_shift):
                # Write employees for the current shift
                row = self.start_row + shift * (self.employees_per_shift + 1) + emp_idx
                # Write employee name in column A
                sheet.cell(row=row, column=1).value = self.employee_names[employee_index]

                # Current shift
                current_shift = shift_pattern

                # Fill schedule: 'P' for weekdays, empty for weekends
                j = 0

                # Init starting column
                start_column = None
                ot_columns = []
                
                for i in range(len(full_weeks)):
                    # print(row, i+2, full_weeks['Weekday'].iloc[i])
                    # Determine weekday (0=Monday, 6=Sunday)
                    if full_weeks['Weekday'].iloc[i] in ("Saturday", "Sunday", "Sat", "Sun"):
                        sheet.cell(row=row, column=j+i+2).value = None

                        # Fixed: Data validation (j) need to be added before skipping to 3
                        data_val.add(sheet.cell(row=row, column=j+i+2))

                        if full_weeks['Weekday'].iloc[i] in ("Sunday", "Sun"):
                            current_shift += 1
                            end_column = colnum_alphabet(j+i+2)
                            # Shift
                            sheet.cell(row=row, column=j+i+3).value =\
                                (f'=COUNTIF({start_column}{row}:{end_column}{row},"P")*8'
                                 f'+ COUNTIF({start_column}{row}:{end_column}{row},"HD")*6')
                            # sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
                            sheet.cell(row=row, column=j+i+3).alignment = Alignment(horizontal='center', vertical='center')
                            # OT
                            sheet.cell(row=row, column=j+i+4).value =\
                                (f'=COUNTIF({start_column}{row}:{end_column}{row},"12")*8'
                                 f'+ COUNTIF({start_column}{row}:{end_column}{row},"T")*6')
                            sheet.cell(row=row, column=j+i+4).alignment = Alignment(horizontal='center', vertical='center')
                            # Append to unique ot_columns
                            ot_columns.append(f"{colnum_alphabet(j+i+4)}{row}")

                            # Total
                            sheet.cell(row=row, column=j+i+5).value =\
                                f'={colnum_alphabet(j+i+3)}{row} + {colnum_alphabet(j+i+4)}{row}'
                            sheet.cell(row=row, column=j+i+5).alignment = Alignment(horizontal='center', vertical='center')
                            j += 3
                    else:
                        # Monday to Friday
                        sheet.cell(row=row, column=j+i+2).value = "P"

                        if full_weeks['Weekday'].iloc[i] in ('Monday', 'Mon'):
                            # Mark starting column on Monday
                            start_column = colnum_alphabet(j+i+2)

                        if current_shift % self.max_shifts != 0:
                            sheet.cell(row=row, column=j+i+2).fill = pattern_fill3
                        data_val.add(sheet.cell(row=row, column=j+i+2))
                employee_index += 1

                # Insert tally
                self.insert_summary_formulas(sheet, full_weeks, row, j, ot_columns)

                # Empty row is handled by the row calculation (shift * (employees_per_shift + 1))

            # Shift pattern change
            shift_pattern += 1

    def insert_summary_formulas(self, sheet, full_weeks, row, j, ot_columns):
        """
        Inserts a block of summary formulas into the row of an employee schedule.

        Args:
            sheet (Worksheet): The openpyxl worksheet.
            full_weeks (DataFrame): The DataFrame containing the schedule dates.
            row (int): The current row number (employee row) to insert formulas into.
            j (int): The starting column offset for the entire schedule block.
            ot_columns (list): A list of Excel cell references for OT columns (e.g., ['D4', 'G4']).
        """
        # Calculate the end column letter for the COUNTIF range.
        # It is the column letter of the last day of the schedule block.
        # The schedule data starts at column j+1, so the last column is j + len(full_weeks).
        last_day_col = colnum_alphabet(j+len(full_weeks))

        # Fill-up the column width top headers
        title = ["P", "OT", "HD", "MC", "T", "AL"]
        for col_offset in range(6):
            col = j + len(full_weeks) + col_offset + 2
            col_letter = colnum_alphabet(col)
            sheet.column_dimensions[col_letter].width = 5
            sheet.cell(row=3, column=col).value = title[col_offset]

            # Top 3 rows pattern fill
            for i in range(1, 4):
                sheet.cell(row=i, column=col).fill = pattern_fill
                sheet.cell(row=i, column=col).fill = pattern_fill
                sheet.cell(row=i, column=col).fill = pattern_fill

        # Define the range for counting scheduled days (e.g., B4:AE4)
        count_range = f"B{row}:{last_day_col}{row}"

        # 1. Define all the required summary calculations and their column offsets
        #    Format: {offset: formula_string_fragment}
        summary_formulas = {
            2: f'=COUNTIF({count_range},"=P")*8 + COUNTIF({count_range},"=HD")*6',  # Total Hours
            4: f'=COUNTIF({count_range},"=HD")',  # Half Days
            5: f'=COUNTIF({count_range},"=MC")',  # Medical Cert (MC)
            6: f'=COUNTIF({count_range},"=T")',  # Training (T)
            7: f'=COUNTIF({count_range},"=AL")'  # Annual Leave (AL)
        }

        # The OT total has a unique formula that joins a list of cell references.
        ot_formula = "={}".format("+".join(map(str, ot_columns)))

        # 2. Insert all COUNTIF formulas
        # The starting column for the summary block is j + len(full_weeks) + 2
        summary_start_col = j + len(full_weeks)

        for offset, formula in summary_formulas.items():
            target_col = summary_start_col + offset
            col_letter = colnum_alphabet(target_col)
            sheet[f"{col_letter}{row}"].value = formula

        # 3. Insert the unique OT total formula (at offset 3)
        ot_col_letter = colnum_alphabet(summary_start_col + 3)
        sheet[f"{ot_col_letter}{row}"].value = ot_formula

    def build_iso_calendar(self, year: int):
        # start_date = f"=DATE({iso_year_start(year).year}, {iso_year_start(year).month}, {iso_year_start(year).day})"
        # print(start_date)

        # 1) compute start/end dates dynamically…
        start = iso_year_start(year)
        end = iso_year_end(year)

        # 2) generate the full date range and ISO metadata
        all_days = pd.date_range(start=start, end=end, freq='D')
        df = pd.DataFrame({'Date': all_days})
        iso = df['Date'].dt.isocalendar()
        df['ISO_Year'] = iso.year
        df['ISO_Week'] = iso.week
        # df['Weekday'] = df['Date'].dt.strftime('%W')
        df['Weekday'] = df['Date'].dt.strftime('%a')

        # 3) keep only the target ISO year
        df = df[df['ISO_Year'] == year]

        # 4) **exclude Week 53** if it exists**
        df = df[df['ISO_Week'] <= 52]

        # 5) map each ISO week to the month of its latest date
        grp = df.groupby('ISO_Week')['Date']

        # default: map each week to the month of its last date
        week_max_month = grp.max().dt.month

        # monday anchor month (first date in each group)
        week_start_month = grp.min().dt.month

        # identify the last ISO week number in your year
        last_week = df['ISO_Week'].max()

        # apply override: final week uses the Monday‐anchor month
        week_month = week_max_month.to_dict()
        week_month[last_week] = week_start_month[last_week]

       # 6) build the workbook as before…
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

        # All sheets will be accumulated here
        sheets: list[Worksheet] = []

        # Test the first week
        # # for m in range(1, 2):
        for m in range(1, 13):
            weeks = [wk for wk, mon in week_month.items() if mon == m]
            month_df = df[df['ISO_Week'].isin(weeks)]
            full_weeks: pd.DataFrame
            full_weeks = month_df.groupby('ISO_Week').filter(lambda g: len(g) == 7)
            if full_weeks.empty:
                continue

            sheet: Worksheet = self.wb.create_sheet(title=date(year, m, 1).strftime('%B'))
            sheets.append(sheet)

            i = 0
            sheet.cell(row=1, column=i+1).value = "Workweek"
            sheet.cell(row=1, column=i+1).fill = pattern_fill2
            sheet.cell(row=2, column=i+1).value = "Weekday"
            sheet.cell(row=2, column=i+1).fill = pattern_fill2
            sheet.cell(row=3, column=i+1).value = "Employee \\ Day"
            sheet.cell(row=3, column=i+1).fill = pattern_fill2
            sheet.column_dimensions['A'].width = 20

            j = 0
            for index, row in full_weeks.iterrows():
                # print(f"Index: {index}, date  {row['Date']}, week {row['ISO_Week']} weekday {row['Weekday']}")
                sheet.column_dimensions[colnum_alphabet(j+i+2)].width = 5
                sheet.cell(row=1, column=j+i+2).value = row['ISO_Week']
                sheet.cell(row=1, column=j+i+2).fill = pattern_fill
                sheet.cell(row=2, column=j+i+2).value = row['Weekday']
                sheet.cell(row=2, column=j+i+2).fill = pattern_fill
                sheet.cell(row=3, column=j+i+2).value = row['Date'].day
                sheet.cell(row=3, column=j+i+2).fill = pattern_fill
                if full_weeks['Weekday'].iloc[i] in ("Sunday", "Sun"):
                    consume_shift(sheet, i, j)
                    j += 3
                i += 1

            self.add_sample_employees(sheet, full_weeks)
            self.add_public_holidays(sheet, full_weeks)

        # All sheets accumulated here to *plural* sheets.
        self.add_public_holidays2(sheets, df)
        outfile = "initial_workweek.xlsx"
        print(f"Writing output file {outfile}...")
        self.wb.save(f"{outfile}")

        print(f"✅ ISO {year} calendar saved (Weeks 1–52 only).")

# # Bootstrap default testing
# builder = CalendarBuilder()
# builder.build_iso_calendar(2025)

# TODO Consecutive two month working days, end of March to beginning of April
# TODO Assuming Shift is mandatory
# TODO Looks like the replacement holiday has been rescheduled automatically by Holidays Python module
# TODO Support for adds weekly parameters such as Shift, OT hours
#  and Monthly accumulation (P, OT, HD, MC, T and AL)
# TODO Complete the Legends
# TODO Wizard to generate spreadsheet
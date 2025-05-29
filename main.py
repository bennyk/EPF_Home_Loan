import math
import csv
from io import StringIO

from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_COMMA_SEPARATED1, FORMAT_NUMBER
from openpyxl.workbook import Workbook
from openpyxl.worksheet import worksheet

# Loan and EPF parameters
initial_loan = 460000  # RM 460,000
interest_rate = 0.042  # 4.2% annual interest rate
tenure_months = 264  # 264 months (22 years)
epf_dividend_rate = 0.055  # 5.5% average annual dividend rate (conservative)
advanced_payment = None


# Function to calculate monthly loan payment
def calculate_monthly_payment(principal, annual_rate, months):
    monthly_rate = annual_rate / 12
    payment = principal * (monthly_rate * (1 + monthly_rate) ** months) / ((1 + monthly_rate) ** months - 1)
    return payment


# Function to generate amortization schedule
def generate_amortization_schedule(principal, drawdown, annual_rate, months, output_file):
    monthly_rate = annual_rate / 12
    monthly_payment = calculate_monthly_payment(principal, annual_rate, months)
    # Drawdown to reduce balance
    balance = principal - drawdown
    schedule = []

    for month in range(1, months + 1):
        interest = balance * monthly_rate
        principal_paid = monthly_payment - interest
        balance -= principal_paid
        if balance < 0:
            balance = 0
        schedule.append({
            'Month': month,
            'Payment': monthly_payment,
            'Interest': interest,
            'Principal': principal_paid,
            'Balance': balance
        })

    # Simulate writing to CSV
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=['Month', 'Payment', 'Interest', 'Principal', 'Balance'])
    writer.writeheader()
    for row in schedule:
        writer.writerow({
            'Month': row['Month'],
            'Payment': f"{row['Payment']:.2f}",
            'Interest': f"{row['Interest']:.2f}",
            'Principal': f"{row['Principal']:.2f}",
            'Balance': f"{row['Balance']:.2f}"
        })

    return schedule, output.getvalue()


# Function to print concise amortization schedule
def print_concise_schedule(schedule, title):
    print(f"\n{title}")
    print("Month | Payment | Interest | Principal | Balance")
    print("-" * 50)

    # Find the last row with a positive balance
    last_index = next((i for i, row in enumerate(schedule) if row['Balance'] <= 0), len(schedule)) - 1

    # Get first 3 and last 3 rows (without overlap)
    first_rows = schedule[:3]
    last_rows = schedule[max(2, last_index - 2):last_index + 1] if last_index >= 2 else []

    # Combine and remove duplicates (in case of overlap)
    shown_months = set()
    for row in first_rows + last_rows:
        if row['Month'] in shown_months:
            continue
        shown_months.add(row['Month'])
        print(f"{row['Month']:>5} | RM {row['Payment']:>7.2f} | RM {row['Interest']:>7.2f}"
              f" | RM {row['Principal']:>8.2f} | RM {row['Balance']:>9.2f}")
        if row['Balance'] <= 0.0:
            break
    print("\nCompleted {:.2f} years later".format(last_rows[-1]['Month']/12))


# Function to calculate EPF growth with compound interest
def calculate_epf_growth(principal, annual_rate, years):
    future_value = principal * (1 + annual_rate) ** years
    return future_value


def print_main(principal, interest_rate, months, total_interest, advanced_payment ):
    print(f"Principal: RM {principal:.2f}")
    print(f"Interest rate: {interest_rate*100:.2f} %")
    print(f"Number of month: {months:.2f}")

    year = months/12
    print(f"Years: {year:.2f}")

    # 460000 * 0.0035 * 1.0035 ** 264 / (1.0035 ** 264–1)
    rate = interest_rate/12
    monthly_payment = principal * rate * (1+rate)**months/((1+rate)**months-1)
    print(f"Monthly Payment: RM {monthly_payment:.2f}")
    print(f"Total Interest: RM {total_interest:.2f}")
    print(f"Repayment: RM {monthly_payment*months:.2f}")
    print(f"Advanced payment: RM {advanced_payment:.2f}")


def write_excel(sheet, principal, interest_rate, months, total_interest, advanced_payment,
                schedule=None):
    sheet.column_dimensions["A"].width = 7
    sheet.column_dimensions["B"].width = 11

    start_column = 3
    row = 1
    sheet.cell(row=row, column=1, value=f"P (Principal)")
    sheet.cell(row=row, column=start_column, value=f"={principal}")
    sheet.cell(row=row, column=start_column).number_format = FORMAT_NUMBER_COMMA_SEPARATED1

    row += 1
    sheet.cell(row=row, column=1, value=f"I (Interest rate)")
    sheet.cell(row=row, column=start_column, value=f"={interest_rate}")
    sheet.cell(row=row, column=start_column).number_format = FORMAT_PERCENTAGE_00

    row += 1
    sheet.cell(row=row, column=1, value=f"Number of month")
    sheet.cell(row=row, column=start_column, value=f"={months}")
    sheet.cell(row=row, column=start_column).number_format = FORMAT_NUMBER

    row += 1
    sheet.cell(row=row, column=1, value=f"Number of years")
    sheet.cell(row=row, column=start_column, value='=COUNTIF(C11:C1032,">1")/12')
    sheet.cell(row=row, column=start_column).number_format = FORMAT_NUMBER

    row += 1
    rate = interest_rate / 12
    monthly_payment = principal * rate * (1+rate) ** months / ((1+rate) ** months - 1)
    sheet.cell(row=row, column=1, value=f"Monthly payment")
    sheet.cell(row=row, column=start_column, value=f"={monthly_payment}")
    sheet.cell(row=row, column=start_column).number_format = FORMAT_NUMBER_COMMA_SEPARATED1

    row += 1
    sheet.cell(row=row, column=1, value=f"Total interest")
    sheet.cell(row=row, column=start_column, value=f"={total_interest}")
    sheet.cell(row=row, column=start_column).number_format = FORMAT_NUMBER_COMMA_SEPARATED1

    row += 1
    sheet.cell(row=row, column=1, value=f"Repayment")
    sheet.cell(row=row, column=start_column, value=f"={monthly_payment*months}")
    sheet.cell(row=row, column=start_column).number_format = FORMAT_NUMBER_COMMA_SEPARATED1

    row += 1
    sheet.cell(row=row, column=1, value=f"Advanced payment")
    sheet.cell(row=row, column=start_column, value=f"={advanced_payment}")
    sheet.cell(row=row, column=start_column).number_format = FORMAT_NUMBER_COMMA_SEPARATED1

    write_concise_schedule(sheet, schedule)


def write_concise_schedule(sheet, schedule, ):
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["E"].width = 11
    base_index = 10
    sheet.cell(row=base_index, column=1, value="Month")
    sheet.cell(row=base_index, column=2, value="Payment")
    sheet.cell(row=base_index, column=3, value="Interest")
    sheet.cell(row=base_index, column=4, value="Principal")
    sheet.cell(row=base_index, column=5, value="Balance")

    # Combine and remove duplicates (in case of overlap)
    shown_months = set()
    for row in schedule:
        if row['Month'] in shown_months:
            continue
        shown_months.add(row['Month'])
        row_adjust = base_index + row['Month']
        sheet.cell(row=row_adjust, column=1, value=row['Month'])

        sheet.cell(row=row_adjust, column=2, value=row['Payment'])
        sheet.cell(row=row_adjust, column=2).number_format = FORMAT_NUMBER_COMMA_SEPARATED1

        sheet.cell(row=row_adjust, column=3, value=row['Interest'])
        sheet.cell(row=row_adjust, column=3).number_format = FORMAT_NUMBER_COMMA_SEPARATED1

        sheet.cell(row=row_adjust, column=4, value=row['Principal'])
        sheet.cell(row=row_adjust, column=4).number_format = FORMAT_NUMBER_COMMA_SEPARATED1

        sheet.cell(row=row_adjust, column=5, value=row['Balance'])
        sheet.cell(row=row_adjust, column=5).number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        if row['Balance'] <= 0.0:
            break


# Scenario 1: Original loan without prepayment
monthly_payment_original = calculate_monthly_payment(initial_loan, interest_rate, tenure_months)
drawdown = 0
schedule_original, csv_original = generate_amortization_schedule(initial_loan, drawdown, interest_rate, tenure_months,
                                                                 "original_loan_amortization.csv")
total_interest_original = sum(row['Interest'] for row in schedule_original)

wb = Workbook()
del wb['Sheet']

# Print results
print("=== Home Loan vs EPF Analysis with Amortization ===")
sheet = wb.create_sheet('Loan without prepayment')
print("\nScenario 1: Continue loan without prepayment")
print_main(principal=initial_loan, interest_rate=interest_rate, months=tenure_months,
           total_interest=total_interest_original, advanced_payment=drawdown)
print_concise_schedule(schedule_original, "Amortization Schedule (Original Loan)")
write_excel(sheet=sheet, principal=initial_loan, interest_rate=interest_rate, months=tenure_months,
            total_interest=total_interest_original, advanced_payment=drawdown, schedule=schedule_original)

# Scenario 2: Loan with prepayment
drawdown = 210000
schedule_original, csv_original = generate_amortization_schedule(initial_loan, drawdown, interest_rate, tenure_months,
                                                                 "original_loan_amortization.csv")
total_interest_original = sum(row['Interest'] for row in schedule_original)

# Print results
sheet = wb.create_sheet('Loan with prepayment')
print("\nScenario 2: Loan with prepayment")
print_main(principal=initial_loan, interest_rate=interest_rate, months=tenure_months,
           total_interest=total_interest_original, advanced_payment=drawdown)
print_concise_schedule(schedule_original, "Amortization Schedule")
write_excel(sheet=sheet, principal=initial_loan, interest_rate=interest_rate, months=tenure_months,
            total_interest=total_interest_original, advanced_payment=drawdown, schedule=schedule_original)

wb.save("out.xlsx")
